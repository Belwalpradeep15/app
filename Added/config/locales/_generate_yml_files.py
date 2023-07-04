#!/usr/bin/env python3
#
# Reads the Translations.xlsx file and generates yml files.
#
import openpyxl
import pprint
import unittest
import yaml


class TestGenerateYmlFiles(unittest.TestCase):

    def test_generate_yml_files(self):
        wb = openpyxl.load_workbook(filename = 'Translations.xlsx', read_only=True)
        ws = wb['Sheet 1']

        """
        Worksheet dimensions
        Read-only mode relies on applications and libraries that created the file providing correct information about the worksheets, specifically the used part of it, known as the dimensions. Some applications set this incorrectly. You can check the apparent dimensions of a worksheet using ws.calculate_dimension(). If this returns a range that you know is incorrect, say A1:A1 then simply resetting the max_row and max_column attributes should allow you to work with the file:
        """
        ws.max_row = ws.max_column = None
        print("worksheet dimensions:", ws.calculate_dimension(force=True))

        # Sanity checks.
        self.assertEqual(1, ws.min_row)
        self.assertLessEqual(1510, ws.max_row)
        self.assertEqual(1, ws.min_column)
        self.assertEqual(5, ws.max_column)

        # These are the expected headings: key + language code.
        headings = ['key', 'en', 'es', 'it', 'tr']
        # These are the output dicts that are populated and then used to generate the *.yml files.
        dicts = [{'placeholder': ''}, {}, {}, {}, {}]

        num_rows_without_keys = 0
        num_translations = 0
        current_row = 0
        # We start with row 2, which are the headings.
        for row in ws.iter_rows(min_row=2):
            current_row += 1

            if current_row == 1:
                # Headings row.
                for cell in row:
                    print('Found heading:', cell.value)
                    self.assertEqual(headings[cell.column - 1], cell.value)

            else:
                # Data row.
                key_parts = []
                for cell in row:
                    if cell.column > len(headings):
                        break
                    if cell.column == 1:
                        # print('key:', cell.value)
                        if cell.value and len(cell.value) > 0:
                            key_parts = cell.value.split('.')
                        else:
                            num_rows_without_keys += 1
                            break
                    else:
                        if len(key_parts) > 0:
                            d = dicts[cell.column - 1]
                            count = 0
                            for key_part in key_parts:
                                count += 1
                                if count < len(key_parts):
                                    if key_part not in d:
                                        d[key_part] = {}
                                    d = d[key_part]
                                else:
                                    v = cell.value
                                    if v is None:
                                        v = ''
                                    v = v.replace('%{NEWLINE}', '\n').replace('%{ESCAPED-NEWLINE}', '\\n')
                                    d[key_part] = v
                                    num_translations += 1

        # pprint.pprint(dicts)
        print('num_rows:', current_row)
        print('num_rows_without_keys:', num_rows_without_keys)
        print('num_translations:', num_translations)

        # Sanity check: num of translations must be 4 * (num of rows processed - 1 (header row) - num of rows without keys).
        self.assertEqual(num_translations, 4 * (current_row - 1 - num_rows_without_keys))

        for h in headings[1:]:
            with open(h + '.yml', 'wb') as f:
                d = {h: dicts[headings.index(h)]}
                yaml.dump(d, stream=f, encoding='utf-8', allow_unicode=True, default_flow_style=False)
                print('Wrote file:', f)


if __name__ == '__main__':
    unittest.main()

